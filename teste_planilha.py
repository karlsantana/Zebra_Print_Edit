from openpyxl import Workbook
wb = Workbook()
# grab the active worksheet
ws = wb.active
# Data can be assigned directly to cells
ws['A1'] = "SERIAL"

# Rows can also be appended
indice = 1
ws.append([1, 2, 7])
while(indice < 10):
    caminho = 'C:/Users/Hardware/Desktop/GRAVADOR/site-packages/REGISTRO/teste.xlsx'
    arquivo_excel = load_workbook(caminho)
    if(arquivo.excel['A%d' % (indice,)]!=0):
    value = input("digite agora um numero")
    ws.cell(row=2, column=indice, value=value)
    indice = indice + 1


#ws.append([1, 2, 7])
#ws.cell(row=2, column=2, value=2)
#ws.cell(row=2, column=3, value=7)
#ws.cell(row=2, column=4, value=257)
# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()
# Save the file
wb.save("C:/Users/Hardware/Desktop/GRAVADOR/site-packages/REGISTRO/teste.xlsx")
#C:\Users\Hardware\Desktop\GRAVADOR\site-packages\REGISTRO